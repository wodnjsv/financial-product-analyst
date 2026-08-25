from __future__ import annotations

from collections.abc import Iterable, Sequence
from pathlib import Path

from openpyxl import Workbook


def write_data_workbook(
    path: Path,
    *,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
    sheet_name: str = "datarows",
) -> Path:
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(tuple(headers))
    for row in rows:
        worksheet.append(tuple(row))
    workbook.save(path)
    return path


def write_schema_workbook(
    path: Path,
    *,
    headers: Sequence[str],
    sheet_name: str = "Sheet1_Schema",
) -> Path:
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(
        ("순번", "컬럼명", "데이터타입", "Nullable", "컬럼코멘트")
    )
    for index, header in enumerate(headers, start=1):
        worksheet.append((index, header, "text", "YES", ""))
    workbook.save(path)
    return path
