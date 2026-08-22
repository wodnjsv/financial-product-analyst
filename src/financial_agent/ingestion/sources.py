from __future__ import annotations

import hashlib
import tempfile
from collections.abc import Iterator, Mapping
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook

from .models import SourceSpec


class SourceVerificationError(RuntimeError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


class ObjectDownloadClient(Protocol):
    def download_file(self, bucket: str, key: str, destination: str) -> None: ...


class ObjectUploadClient(ObjectDownloadClient, Protocol):
    def upload_file(self, source: str, bucket: str, key: str) -> None: ...


def sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    try:
        with path.open("rb") as source:
            for chunk in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(chunk)
    except OSError:
        raise SourceVerificationError(
            "SOURCE_READ_FAILED", "local source could not be read"
        ) from None
    return digest.hexdigest()


def verify_local_source(path: Path, expected_sha256: str | None = None) -> str:
    actual_sha256 = sha256_path(path)
    if expected_sha256 is not None and actual_sha256 != expected_sha256:
        raise SourceVerificationError(
            "SOURCE_CHECKSUM_MISMATCH",
            "local source checksum differs from expected",
        )
    return actual_sha256


def download_verified_object(
    client: ObjectDownloadClient,
    *,
    bucket: str,
    key: str,
    expected_sha256: str,
    destination: Path,
) -> Path:
    temporary_path: Path | None = None
    try:
        destination.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(
            prefix=".organizer-source-",
            dir=destination.parent,
            delete=False,
        ) as temporary:
            temporary_path = Path(temporary.name)
        client.download_file(bucket, key, str(temporary_path))
    except Exception:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)
        raise SourceVerificationError(
            "OBJECT_DOWNLOAD_FAILED", "object download failed"
        ) from None

    actual_sha256 = sha256_path(temporary_path)
    if actual_sha256 != expected_sha256:
        temporary_path.unlink(missing_ok=True)
        raise SourceVerificationError(
            "OBJECT_CHECKSUM_MISMATCH",
            "downloaded object checksum differs from expected",
        )
    try:
        temporary_path.replace(destination)
    except OSError:
        temporary_path.unlink(missing_ok=True)
        raise SourceVerificationError(
            "OBJECT_DOWNLOAD_FAILED", "object download failed"
        ) from None
    return destination


def upload_verified_object(
    client: ObjectUploadClient,
    *,
    bucket: str,
    key: str,
    source: Path,
    expected_sha256: str,
) -> str:
    if sha256_path(source) != expected_sha256:
        raise SourceVerificationError(
            "OBJECT_UPLOAD_SOURCE_CHECKSUM_MISMATCH",
            "upload source checksum differs from expected",
        ) from None

    verification_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=".object-upload-verification-",
            dir=source.parent,
            delete=False,
        ) as temporary:
            verification_path = Path(temporary.name)
        client.upload_file(str(source), bucket, key)
        client.download_file(bucket, key, str(verification_path))
        if sha256_path(verification_path) != expected_sha256:
            raise SourceVerificationError(
                "OBJECT_UPLOAD_VERIFICATION_FAILED",
                "uploaded object checksum differs from expected",
            )
    except SourceVerificationError:
        raise
    except Exception:
        raise SourceVerificationError(
            "OBJECT_UPLOAD_FAILED", "object upload failed"
        ) from None
    finally:
        if verification_path is not None:
            verification_path.unlink(missing_ok=True)
    return expected_sha256


def verify_schema_header(path: Path, spec: SourceSpec) -> tuple[str, ...]:
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if spec.schema_sheet_name not in workbook.sheetnames:
            raise SourceVerificationError(
                "SOURCE_SCHEMA_SHEET_MISSING",
                "approved schema sheet is missing",
            )
        worksheet = workbook[spec.schema_sheet_name]
        columns = tuple(
            row[0]
            for row in worksheet.iter_rows(min_row=3, values_only=True)
            if row[0] is not None
        )
        if columns != spec.expected_columns:
            raise SourceVerificationError(
                "SOURCE_SCHEMA_MISMATCH",
                "schema columns differ from the approved source",
            )
        return spec.expected_columns
    except SourceVerificationError:
        raise
    except Exception:
        raise SourceVerificationError(
            "SOURCE_WORKBOOK_READ_FAILED", "source workbook could not be read"
        ) from None
    finally:
        if workbook is not None:
            workbook.close()


def iter_workbook_rows(
    path: Path, spec: SourceSpec
) -> Iterator[Mapping[str, object]]:
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if spec.data_sheet_name not in workbook.sheetnames:
            raise SourceVerificationError(
                "SOURCE_DATA_SHEET_MISSING", "approved data sheet is missing"
            )
        worksheet = workbook[spec.data_sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        header = tuple(header_row) if header_row is not None else ()
        if header != spec.expected_columns:
            raise SourceVerificationError(
                "SOURCE_HEADER_MISMATCH",
                "workbook header differs from the approved schema",
            )

        row_count = 0
        for values in rows:
            row_count += 1
            if len(values) > len(spec.expected_columns):
                raise SourceVerificationError(
                    "SOURCE_ROW_WIDTH_MISMATCH",
                    "workbook row is wider than the approved schema",
                )
            padded_values = values + (None,) * (
                len(spec.expected_columns) - len(values)
            )
            yield dict(
                zip(spec.expected_columns, padded_values, strict=True)
            )
        if row_count != spec.expected_row_count:
            raise SourceVerificationError(
                "SOURCE_ROW_COUNT_MISMATCH",
                "workbook row count differs from the approved source",
            )
    except SourceVerificationError:
        raise
    except Exception:
        raise SourceVerificationError(
            "SOURCE_WORKBOOK_READ_FAILED", "source workbook could not be read"
        ) from None
    finally:
        if workbook is not None:
            workbook.close()
